import discord
import datetime
from random import *
import requests
from bs4 import BeautifulSoup as bs
import sys
import openpyxl
import asyncio
import time
import os



client = discord.Client()




@client.event
async def on_ready():
    print("Run_Bot")
    print(client.user.name)
    print(client.user.id)
    print("-------------------------\n")
    await client.change_presence(game=discord.Game(name='/?을(를) 입력하여 명령어 확인          ', type=0))




@client.event
async def on_message(message):
    if message.channel.id != '542335761962237972' and message.channel.id != '542335831675764736':
        now = datetime.datetime.now()



    if message.content == "/게임" or message.content == "/ㄱㅇ" or message.content == "/겜" or message.content == "/ㄱ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        embed = discord.Embed(title="Sleep 상태에서 제한된 명령어입니다.", description="나중에 다시 시도해 주세요.", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(message.channel, embed=embed)
        


    if message.content == "/게임 가위바위보" or message.content == "/ㄱㅇ ㄱㅇㅂㅇㅂ" or message.content == "/겜 가위바위보" or message.content == "/ㄱ ㄱㅇㅂㅇㅂ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        embed = discord.Embed(title="Sleep 상태에서 제한된 명령어입니다.", description="나중에 다시 시도해 주세요.", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(message.channel, embed=embed)


    if message.content == "/경고기록" or message.content == "/ㄱㄱㄱㄹ" or message.content == "/경고조회" or message.content == "/ㄱㄱㅈㅎ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        file = openpyxl.load_workbook("warning_list.xlsx")
        sheet = file.active
        for i in range(1, 129):
            if str(sheet["B" + str(i)].value) == str(message.author.id):
                embed = discord.Embed(title="경고 기록" , description="현재 <@" + message.author.id + ">님의 경고 횟수는 " + str(sheet["C" + str(i)].value) + "회 입니다.", color=0x00ff00)
                embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
                await client.send_message(message.channel, embed=embed)
                break
            if str(sheet["B" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(message.author.name)
                sheet["B" + str(i)].value = str(message.author.id)
                sheet["C" + str(i)].value = "0"
                embed = discord.Embed(title="경고 기록" , description="현재 <@" + message.author.id + ">님의 경고 횟수는 " + str(sheet["C" + str(i)].value) + "회 입니다.", color=0x00ff00)
                embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
                await client.send_message(message.channel, embed=embed)
                break
        file.save("warning_list.xlsx")

    elif message.content.startswith('/경고기록 ') or message.content.startswith('/ㄱㄱㄱㄹ ') or message.content.startswith('/경고조회 ') or message.content.startswith('/ㄱㄱㅈㅎ '):
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        warnhistory = message.content[6:]
        file = openpyxl.load_workbook("warning_list.xlsx")
        sheet = file.active
        for i in range(1, 129):
            if str(sheet["B" + str(i)].value) == str(warnhistory):
                embed = discord.Embed(title="경고 기록" , description="현재 <@" + warnhistory + ">님의 경고 횟수는 " + str(sheet["C" + str(i)].value) + "회 입니다.", color=0x00ff00)
                embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
                await client.send_message(message.channel, embed=embed)
                break
            if str(sheet["B" + str(i)].value) == "-":
                embed = discord.Embed(title="오류" , description="<@" + warnhistory + ">님의 경고 기록이 없거나 찾을 수 없습니다.", color=0x00ff00)
                embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
                await client.send_message(message.channel, embed=embed)
                break
        file.save("warning_list.xlsx")


        
    if message.content == "/변경사항" or message.content == "/ㅂㄳㅎ" or message.content == "/ㅂㄱㅅㅎ" or message.content == "/체인지로그" or message.content == "/ㅊㅇㅈㄺ" or message.content == "/ㅊㅇㅈㄹㄱ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        embed = discord.Embed(title="1.4.0 변경 사항", description="``/경고기록`` 및 ``/경고조회`` 명령어를 추가하였습니다.", color=0x00ff00)
        await client.send_message(message.channel, embed=embed)



    if message.content == "/게임 주사위" or message.content == "/ㄱㅇ ㅈㅅㅇ" or message.content == "/겜 주사위" or message.content == "/ㄱ ㅈㅅㅇ" or message.content == "/게임 다이스" or message.content == "/겜 다이스" or message.content == "/ㄱㅇ ㄷㅇㅅ" or message.content == "/ㄱ ㄷㅇㅅ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        embed = discord.Embed(title="Sleep 상태에서 제한된 명령어입니다.", description="나중에 다시 시도해 주세요.", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(message.channel, embed=embed)



    if message.content == "/엔타로아둔" or message.content == "/ㅇㅌㄹㅇㄷ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        embed = discord.Embed(title="En Taro Adun!", description="엙 타로 아둔!", color=0x00ff00)
        await client.send_message(message.channel, embed=embed)



    if message.content == "/도움" or message.content == "/ㄷㅇ" or message.content == "/헬프" or message.content == "/핼프" or message.content == "/ㅎㅍ" or message.content == "/커맨드" or message.content == "/ㅋㅁㄷ" or message.content == "/명령어" or message.content == "/ㅁㄹㅇ" or message.content == "/?":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        embed = discord.Embed(title="사용 가능한 명령어", description="\n"
                                                              "``/?`` ``/도움`` ``/헬프`` ``/핼프`` ``/커맨드`` ``/명령어``\n사용 가능한 명령어 목록을 보여줍니다.\n\n"
                                                              "``/개발용도움`` ``/개발용헬프`` ``/개발용핼프``\n개발용 명령어 목록을 보여줍니다.\n\n"
                                                              "``/검색 (텍스트)`` ``/서치 (텍스트)``\nGoogle에서 이미지를 불러옵니다. (텍스트)에 공백이 없어야 정상 작동합니다.\n\n"
                                                              "``/게임`` ``/겜``\n~~현재 자신의 게임 점수를 보여줍니다.~~ **Sleep 상태에서 사용 불가**\n\n"
                                                              "``/게임 가위바위보`` ``/겜 가위바위보``\n~~가위바위보 게임을 진행합니다.~~ **Sleep 상태에서 사용 불가**\n\n"
                                                              "``/게임 주사위`` ``/게임 다이스`` ``/겜 주사위`` ``/겜 다이스``\n~~주사위를 굴립니다.~~ **Sleep 상태에서 사용 불가**\n\n"
                                                              "``/경고기록`` ``/경고조회``\n자신의 경고 기록을 조회합니다.\n\n"
                                                              "``/경고기록 (유저 ID)`` ``/경고조회 (유저ID)``\n(유저 ID)의 경고 기록을 조회합니다.\n\n"
                                                              "``/명언``\n명언을 남깁니다.\n\n"
                                                              "``/변경사항`` ``/체인지로그``\n최근 변경 사항을 보여줍니다.\n\n"
                                                              "``/봇소개``\n봇의 소개를 보여줍니다.\n\n"
                                                              "``/시간`` ``/타임``\n현재 시간을 보여줍니다.\n\n"
                                                              "``/언급`` ``/멘션``\n명령어를 사용한 유저에게 언급합니다.\n\n"
                                                              "``/엔타로아둔``\n인사를 합니다.\n\n"
                                                              "```ini\n[ 이 명령어들은 초성으로도 사용 가능합니다. ]\n```\n\n"
                                                              , color=0x00ff00)
        await client.send_message(message.channel, embed=embed)



    if message.content == "/개발용도움" or message.content == "/ㄱㅂㅇㄷㅇ" or message.content == "/개발용헬프" or message.content == "/개발용핼프" or message.content == "/ㄱㅂㅇㅎㅍ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        embed = discord.Embed(title="개발용 명령어", description="\n"
                                                              "``/r (텍스트)``\n(텍스트)를 봇이 따라합니다. 세부 사항은 소스 코드를 참조하세요.\n\n"
                                                              "``/v``\n현재 버전을 확인합니다.\n\n"
                                                              "``/pn``\n패치 노트를 보여줍니다.\n\n"
                                                              "``/pn f``\n패치 노트 - 세부 사항을 보여줍니다.\n\n"
                                                              "``/n``\n공지사항을 게시합니다. 세부 사항은 소스 코드를 참조하세요.\n\n"
                                                              "```ini\n[ 개발용 명령어는 봇 개발자만 사용할 수 있습니다. ]\n```\n\n"
                                                              , color=0x00ff00)
        await client.send_message(message.channel, embed=embed)



    if message.content == "/언급" or message.content == "/ㅇㄱ" or message.content == "/멘션" or message.content == "/ㅁㅅ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        embed = discord.Embed(title="내가 너를 멘션하겠노라.", description="<@" + message.author.id + ">", color=0x00ff00)
        await client.send_message(message.channel, embed=embed)



    if message.content == "/시간" or message.content == "/ㅅㄱ" or message.content == "/타임" or message.content == "/ㅌㅇ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        embed = discord.Embed(title="현재 시각", description="", color=0x00ff00)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(message.channel, embed=embed)



    if message.content.split(" ")[0] == "/검색" or message.content.split(" ")[0] == "/ㄱㅅ" or message.content.split(" ")[0] == "/ㄳ" or message.content.split(" ")[0] == "/서치" or message.content.split(" ")[0] == "/ㅅㅊ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        group = message.content.split(" ")[1]
        google_data = requests.get("https://google.co.kr/search?q=" + group + "&dcr=0&source=lnms&tbm=isch&sa=X")
        soup = bs(google_data.text, "html.parser")
        imgs = soup.find_all("img")
        await client.send_message(message.channel, imgs[1]['src'])
        del group, google_data, soup, imgs



    if message.content == "/명언" or message.content == "/ㅁㅇ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        embed = discord.Embed(title="알림", description="9.9148542초 이내에 채팅을 입력해주세요.", color=0x00ff00)
        await client.send_message(message.channel, embed=embed)
        wisesay = await client.wait_for_message(timeout=10, author=message.author)
        if wisesay is None:
            embed = discord.Embed(title="삐빅-", description="시간초과!", color=0x00ff00)
            await client.send_message(message.channel, embed=embed)
            return
        else:
            embed = discord.Embed(title=wisesay.content, description="-<@" + message.author.id + ">", color=0x00ff00)
            await client.send_message(message.channel, embed=embed)
            embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
            embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
            await client.send_message(client.get_channel('542335831675764736'), embed=embed)
            file = open('log.txt', 'a', encoding='UTF-8')
            file.write('\n   \n[Channel]: %s(%s) | [Author]: %s(%s) | [Message]:\n%s' % (message.channel, message.channel.id, message.author.name, message.author.id, message.content))



    if message.content == "/봇소개" or message.content == "/ㅂㅅㄱ" or message.content == "/ㅄㄱ":
        embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
        embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
        await client.send_message(client.get_channel('542335831675764736'), embed=embed)
        file = open('log.txt', 'a', encoding='UTF-8')
        file.write('\n   \n[Channel]: %s(%s) | [Author]: %s(%s) | [Message]:\n%s' % (message.channel, message.channel.id, message.author.name, message.author.id, message.content))
        embed = discord.Embed(title="ㅎㅇ 내가 누구냐고?", description="난 <@540110454077390850>임.", color=0x00ff00)
        await client.send_message(message.channel, embed=embed)
        


    if message.author.id == '150577293981515776' or message.author.id == '311791989681291264':
        if message.content == ".vote1":
            embed = discord.Embed(title="1차 투표가 시작되었습니다." , description="재판에 올릴 플레이어를 지목하여 주세요.", color=0x8000ff)
            embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
            await client.send_message(client.get_channel('548548318830133278'), embed=embed)
        if message.content.startswith('.argument '):
            embed = discord.Embed(title="재판에 올려질 플레이어가 결정되었습니다." , description="재판에 올려진 <@" + message.content[10:] + ">님은 변론을 해주세요.", color=0x8000ff)
            embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
            await client.send_message(client.get_channel('548548318830133278'), embed=embed)
        if message.content == ".vote2":
            embed = discord.Embed(title="2차 투표가 시작되었습니다." , description="재판에 올려진 플레이어에 대해 찬반 투표를 해주세요.\n투표는 **찬성**, **반대**, **기권**을 하실 수 있습니다.", color=0x8000ff)
            embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
            member = discord.utils.get(client.get_all_members())
            vote2 = await client.send_message(client.get_channel('548548318830133278'), embed=embed)
            #await client.add_reaction(vote2, '👍')
            #await client.add_reaction(vote2, '👎')
        if message.content.startswith('.daytime '):
            embed = discord.Embed(title=message.content[9:] + "번째 낮이 되었습니다." , description="이제 생존한 플레이어는 <#548548289222410241> 채널을 사용할 수 있습니다.", color=0x0691fa)
            embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
            await client.send_message(client.get_channel('548548289222410241'), embed=embed)
        if message.content.startswith('.night '):
            embed = discord.Embed(title=message.content[7:] + "번째 밤이 되었습니다." , description="이제 생존한 플레이어는 <#548548289222410241> 채널을 사용할 수 없습니다.\n(가능한 경우) 자신의 역할 채널에서 능력을 사용해주세요.", color=0x0691fa)
            embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
            await client.send_message(client.get_channel('548548289222410241'), embed=embed)


    if message.author.id == '150577293981515776':

        if message.content.startswith('/r '):
            await client.send_message(message.channel, message.content[3:])

        if message.content.startswith('/rg '):
            await client.send_message(client.get_channel('485893152738377768'), message.content[3:])

        if message.content.startswith('/rb '):
            await client.send_message(client.get_channel('507932999573045259'), message.content[3:])

        if message.content.startswith('/rm '):
            await client.send_message(client.get_channel('548548289222410241'), message.content[3:])

        if message.content == "/v":
            await client.send_message(message.channel, "현재 버전은 ``1.4.0.0000 (r:47-190225)``입니다.")
        
   
    #if message.channel.id != '548548456436990003' and message.channel.id != '548548289222410241' and message.channel.id != '548548318830133278' and message.channel.id != '548571470486568960':
    if message.server.id != '548547278609317910':
        if \
        "10새" in message.content or \
        "10새기" in message.content or \
        "10새리" in message.content or \
        "10세리" in message.content or \
        "10쉐이" in message.content or \
        "10쉑" in message.content or \
        "10쌔" in message.content or \
        "10쌔기" in message.content or \
        "10쎄" in message.content or \
        "10알" in message.content or \
        "10창" in message.content or \
        "10탱" in message.content or \
        "18것" in message.content or \
        "18넘" in message.content or \
        "18년" in message.content or \
        "18노" in message.content or \
        "18놈" in message.content or \
        "18뇬" in message.content or \
        "18럼" in message.content or \
        "18롬" in message.content or \
        "18새" in message.content or \
        "18새끼" in message.content or \
        "18색" in message.content or \
        "18세끼" in message.content or \
        "18세리" in message.content or \
        "18섹" in message.content or \
        "18쉑" in message.content or \
        "18스" in message.content or \
        "18아" in message.content or \
        "c파" in message.content or \
        "c팔" in message.content or \
        "ㄱㅐ" in message.content or \
        "ㄱ ㅐ" in message.content or \
        "ㄱ  ㅐ" in message.content or \
        "ㄱ   ㅐ" in message.content or \
        "ㄱ        ㅐ" in message.content or \
        "ㄱ         ㅐ" in message.content or \
        "ㄱ          ㅐ" in message.content or \
        "ㄱ           ㅐ" in message.content or \
        "ㄱ                ㅐ" in message.content or \
        "ㄱ                 ㅐ" in message.content or \
        "ㄱ1ㅐ" in message.content or \
        "ㄱ2ㅐ" in message.content or \
        "ㄱ3ㅐ" in message.content or \
        "ㄱ4ㅐ" in message.content or \
        "ㄱ5ㅐ" in message.content or \
        "갈보" in message.content or \
        "갈보년" in message.content or \
        "강아지" in message.content or \
        "같은년" in message.content or \
        "같은뇬" in message.content or \
        "개같은" in message.content or \
        "개구라" in message.content or \
        "개년" in message.content or \
        "개놈" in message.content or \
        "개뇬" in message.content or \
        "개대중" in message.content or \
        "개독" in message.content or \
        "개돼중" in message.content or \
        "개랄" in message.content or \
        "개뻥" in message.content or \
        "개뿔" in message.content or \
        "개새" in message.content or \
        "개새기" in message.content or \
        "개새끼" in message.content or \
        "개새키" in message.content or \
        "개색기" in message.content or \
        "개색끼" in message.content or \
        "개색키" in message.content or \
        "개색히" in message.content or \
        "개섀끼" in message.content or \
        "개세" in message.content or \
        "개세끼" in message.content or \
        "개세이" in message.content or \
        "개소리" in message.content or \
        "개수작" in message.content or \
        "개쉐" in message.content or \
        "개쉐리" in message.content or \
        "개쉐이" in message.content or \
        "개쉑" in message.content or \
        "개쉽" in message.content or \
        "개스끼" in message.content or \
        "개시키" in message.content or \
        "개십새기" in message.content or \
        "개십새끼" in message.content or \
        "개쐑" in message.content or \
        "개씹" in message.content or \
        "개아들" in message.content or \
        "개자슥" in message.content or \
        "개접" in message.content or \
        "개좆" in message.content or \
        "개좌식" in message.content or \
        "개허접" in message.content or \
        "걔새" in message.content or \
        "걔수작" in message.content or \
        "걔시끼" in message.content or \
        "걔시키" in message.content or \
        "걔썌" in message.content or \
        "걸레년" in message.content or \
        "게색기" in message.content or \
        "게색끼" in message.content or \
        "광뇬" in message.content or \
        "구녕" in message.content or \
        "구라" in message.content or \
        "구멍그년" in message.content or \
        "그새끼" in message.content or \
        "ㄲㅏ" in message.content or \
        "ㄲㅑ" in message.content or \
        "ㄲㅣ" in message.content or \
        "냄비" in message.content or \
        "놈현" in message.content or \
        "뇬" in message.content or \
        "눈깔" in message.content or \
        "뉘미럴" in message.content or \
        "니귀미" in message.content or \
        "니기미" in message.content or \
        "니미" in message.content or \
        "니미랄" in message.content or \
        "니미럴" in message.content or \
        "니아배" in message.content or \
        "니아베" in message.content or \
        "니아비" in message.content or \
        "니어매" in message.content or \
        "니어메" in message.content or \
        "니어미" in message.content or \
        "닝기리" in message.content or \
        "닝기미" in message.content or \
        "뎡신" in message.content or \
        "도라이" in message.content or \
        "돈놈" in message.content or \
        "돌아이" in message.content and\
        "돌은놈" in message.content or \
        "되질래" in message.content or \
        "뒈져" in message.content or \
        "뒈져라" in message.content or \
        "뒈진" in message.content or \
        "뒈진다" in message.content or \
        "등신" in message.content or \
        "디져라" in message.content or \
        "디진다" in message.content or \
        "디질래" in message.content or \
        "딩시" in message.content or \
        "따식" in message.content or \
        "때놈" in message.content or \
        "또라이" in message.content or \
        "똘아이" in message.content or \
        "뙤놈" in message.content or \
        "뙨넘" in message.content or \
        "뙨놈" in message.content or \
        "뚜쟁" in message.content or \
        "띠바" in message.content or \
        "띠발" in message.content or \
        "띠불" in message.content or \
        "띠팔" in message.content or \
        "ㅂ ㅅ" in message.content or \
        "ㅂ@ㅅ" in message.content or \
        "ㅂ1ㅅ" in message.content or \
        "ㅂㅅ" in message.content or \
        "바랄년" in message.content or \
        "뱅 신" in message.content or \
        "뱅1신" in message.content or \
        "뱅마" in message.content or \
        "뱅신" in message.content or \
        "벼1엉1신" in message.content or \
        "벼1엉신" in message.content or \
        "벼엉1신" in message.content or \
        "벼엉신" in message.content or \
        "병쉰" in message.content or \
        "병신" in message.content or \
        "병자" in message.content or \
        "보지" in message.content or \
        "부랄" in message.content or \
        "부럴" in message.content or \
        "불알" in message.content or \
        "불할" in message.content or \
        "붕가" in message.content or \
        "뷰웅" in message.content or \
        "븅" in message.content or \
        "븅신" in message.content or \
        "빌어먹" in message.content or \
        "빙시" in message.content or \
        "빙신" in message.content or \
        "빠가" in message.content or \
        "빠구리" in message.content or \
        "빠굴" in message.content or \
        "빠큐" in message.content or \
        "뻐큐" in message.content or \
        "뻑큐" in message.content or \
        "뽁큐" in message.content or \
        "ㅄ" in message.content or \
        "ㅅ ㅂ" in message.content or \
        "ㅅ@ㅂ" in message.content or \
        "ㅅ1ㅂ" in message.content or \
        "ㅅㅂ" in message.content or \
        "ㅅㅂㄹㅁ" in message.content or \
        "ㅅㅐ" in message.content or \
        "ㅅ ㅐ" in message.content or \
        "ㅅ  ㅐ" in message.content or \
        "ㅅ   ㅐ" in message.content or \
        "ㅅ        ㅐ" in message.content or \
        "상넘이" in message.content or \
        "상놈을" in message.content or \
        "상놈의" in message.content or \
        "상놈이" in message.content or \
        "새갸" in message.content or \
        "새꺄" in message.content or \
        "새끼" in message.content or \
        "새새끼" in message.content or \
        "새키색끼" in message.content or \
        "색스" in message.content or \
        "샋" in message.content or \
        "생쑈" in message.content or \
        "섊" in message.content or \
        "세갸" in message.content or \
        "세꺄" in message.content or \
        "세끼" in message.content or \
        "섹스" in message.content or \
        "섻" in message.content or \
        "쇼하네" in message.content or \
        "쉐" in message.content or \
        "쉐기" in message.content or \
        "쉐끼" in message.content or \
        "쉐리" in message.content or \
        "쉐에기" in message.content or \
        "쉐키" in message.content or \
        "쉑" in message.content or \
        "쉣" in message.content or \
        "쉨" in message.content or \
        "쉬발" in message.content or \
        "쉬밸" in message.content or \
        "쉬벌" in message.content or \
        "쉬뻘" in message.content or \
        "쉬펄" in message.content or \
        "쉽알" in message.content or \
        "스패킹" in message.content or \
        "스팽" in message.content or \
        "시 발" in message.content or \
        "시  발" in message.content or \
        "시   발" in message.content or \
        "시        발" in message.content or \
        "시         발" in message.content or \
        "시1발" in message.content or \
        "시12발" in message.content or \
        "시123발" in message.content or \
        "시2발" in message.content or \
        "시3발" in message.content or \
        "시궁창" in message.content or \
        "시끼" in message.content or \
        "시댕" in message.content or \
        "시뎅" in message.content or \
        "시랄" in message.content or \
        "시발" in message.content or \
        "시벌" in message.content or \
        "시부랄" in message.content or \
        "시부럴" in message.content or \
        "시부리" in message.content or \
        "시불" in message.content or \
        "시브랄" in message.content or \
        "시팍" in message.content or \
        "시팔" in message.content or \
        "시펄" in message.content or \
        "심탱" in message.content or \
        "십8" in message.content or \
        "십라" in message.content or \
        "십새" in message.content or \
        "십새끼" in message.content or \
        "십세" in message.content or \
        "십쉐" in message.content or \
        "십쉐이" in message.content or \
        "십스키" in message.content or \
        "십쌔" in message.content or \
        "십창" in message.content or \
        "십탱" in message.content or \
        "싶알" in message.content or \
        "ㅆㅂ" in message.content or \
        "ㅆㅂㄹㅁ" in message.content or \
        "ㅆ앙" in message.content or \
        "ㅆㅍ" in message.content or \
        "ㅆㅣ" in message.content or \
        "싸가지" in message.content or \
        "싹아지" in message.content or \
        "쌉년" in message.content or \
        "쌍넘" in message.content or \
        "쌍년" in message.content or \
        "쌍놈" in message.content or \
        "쌍뇬" in message.content or \
        "쌔끼쌕" in message.content or \
        "쌩쑈" in message.content or \
        "쌴년" in message.content or \
        "썅" in message.content or \
        "썅년" in message.content or \
        "썅놈" in message.content or \
        "썡쇼" in message.content or \
        "써벌" in message.content or \
        "썩을년" in message.content or \
        "썩을놈" in message.content or \
        "쎄꺄" in message.content or \
        "쎄엑" in message.content or \
        "쒸벌" in message.content or \
        "쒸뻘" in message.content or \
        "쒸팔" in message.content or \
        "쒸펄" in message.content or \
        "쓰바" in message.content or \
        "쓰박" in message.content or \
        "쓰발" in message.content or \
        "쓰벌" in message.content or \
        "쓰팔" in message.content or \
        "씁새" in message.content or \
        "씁얼" in message.content or \
        "씌파" in message.content or \
        "씨 발" in message.content or \
        "씨@발" in message.content or \
        "씨1발" in message.content or \
        "씨8" in message.content or \
        "씨끼" in message.content or \
        "씨댕" in message.content or \
        "씨뎅" in message.content or \
        "씨바" in message.content or \
        "씨바랄" in message.content or \
        "씨박" in message.content or \
        "씨발" in message.content or \
        "씨 발" in message.content or \
        "씨  발" in message.content or \
        "씨   발" in message.content or \
        "씨        발" in message.content or \
        "씨         발" in message.content or \
        "씨          발" in message.content or \
        "씨           발" in message.content or \
        "씨                발" in message.content or \
        "씨1발" in message.content or \
        "씨12발" in message.content or \
        "씨123발" in message.content or \
        "씨방" in message.content or \
        "씨방새" in message.content or \
        "씨방세" in message.content or \
        "씨밸" in message.content or \
        "씨뱅" in message.content or \
        "씨벌" in message.content or \
        "씨벨" in message.content or \
        "씨봉" in message.content or \
        "씨봉알" in message.content or \
        "씨부랄" in message.content or \
        "씨부럴" in message.content or \
        "씨부렁" in message.content or \
        "씨부리" in message.content or \
        "씨불" in message.content or \
        "씨붕" in message.content or \
        "씨브랄" in message.content or \
        "씨빠" in message.content or \
        "씨빨" in message.content or \
        "씨뽀랄" in message.content or \
        "씨앙" in message.content or \
        "씨파" in message.content or \
        "씨팍" in message.content or \
        "씨팔" in message.content or \
        "씨펄" in message.content or \
        "씸년" in message.content or \
        "씸뇬" in message.content or \
        "씸새끼" in message.content or \
        "씹같" in message.content or \
        "씹년" in message.content or \
        "씹뇬" in message.content or \
        "씹새" in message.content or \
        "씹새기" in message.content or \
        "씹새끼" in message.content or \
        "씹새리" in message.content or \
        "씹세" in message.content or \
        "씹쉐" in message.content or \
        "씹스키" in message.content or \
        "씹쌔" in message.content or \
        "씹이" in message.content or \
        "씹질" in message.content or \
        "씹창" in message.content or \
        "씹탱" in message.content or \
        "씹퇭" in message.content or \
        "씹팔" in message.content or \
        "씹할" in message.content or \
        "씹헐" in message.content or \
        "아가리" in message.content or \
        "아갈" in message.content or \
        "아갈이" in message.content or \
        "아갈통" in message.content or \
        "아구창" in message.content or \
        "아구통" in message.content or \
        "아굴" in message.content or \
        "얌마" in message.content or \
        "양넘" in message.content or \
        "양년" in message.content or \
        "양놈" in message.content or \
        "엄창" in message.content or \
        "엠병" in message.content or \
        "여물통" in message.content or \
        "염병" in message.content or \
        "엿같" in message.content or \
        "옘병" in message.content or \
        "옘빙" in message.content or \
        "오입" in message.content or \
        "왜년" in message.content or \
        "왜놈" in message.content or \
        "욤병" in message.content or \
        "이년" in message.content or \
        "이새끼" in message.content or \
        "이새키" in message.content or \
        "이스끼" in message.content or \
        "이스키" in message.content or \
        "자지" in message.content or \
        "잡것" in message.content or \
        "잡넘" in message.content or \
        "잡년" in message.content or \
        "잡놈" in message.content or \
        "저년" in message.content or \
        "저새끼" in message.content or \
        "접년" in message.content or \
        "젖밥" in message.content or \
        "조까" in message.content or \
        "조까치" in message.content or \
        "조낸" in message.content or \
        "조또" in message.content or \
        "조랭" in message.content or \
        "조빠" in message.content or \
        "조쟁이" in message.content or \
        "조지냐" in message.content or \
        "조진다" in message.content or \
        "조질래" in message.content or \
        "조찐" in message.content or \
        "존나" in message.content or \
        "존나게" in message.content or \
        "존니" in message.content or \
        "존만" in message.content or \
        "존만한" in message.content or \
        "좀물" in message.content or \
        "좁년" in message.content or \
        "좁밥" in message.content or \
        "좃까" in message.content or \
        "좃또" in message.content or \
        "좃만" in message.content or \
        "좃밥" in message.content or \
        "좃이" in message.content or \
        "좃찐" in message.content or \
        "좆" in message.content or \
        "좌식" in message.content or \
        "주글" in message.content or \
        "주글래" in message.content or \
        "주데이" in message.content or \
        "주뎅" in message.content or \
        "주뎅이" in message.content or \
        "주둥아리" in message.content or \
        "주둥이" in message.content or \
        "주접" in message.content or \
        "주접떨" in message.content or \
        "죽고잡" in message.content or \
        "죽을래" in message.content or \
        "죽통" in message.content or \
        "쥐랄" in message.content or \
        "쥐롤" in message.content or \
        "지랄" in message.content or \
        "지럴" in message.content or \
        "지롤" in message.content or \
        "지미랄" in message.content or \
        "쫍빱" in message.content or \
        "쮸발" in message.content or \
        "쮸밤" in message.content or \
        "쮸뱔" in message.content or \
        "쮸뱜" in message.content or \
        "찌랄" in message.content or \
        "창녀" in message.content or \
        "凸" in message.content or \
        "캐년" in message.content or \
        "캐놈" in message.content or \
        "캐스끼" in message.content or \
        "캐스키" in message.content or \
        "캐시키" in message.content or \
        "탱구" in message.content or \
        "팔럼" in message.content or \
        "퍽유" in message.content or \
        "퍽큐" in message.content or \
        "하아앙.." in message.content or \
        "하앙.." in message.content or \
        "하으읏.." in message.content or \
        "하읏.." in message.content or \
        "할짝" in message.content or \
        "할쨕" in message.content or \
        "핥짝" in message.content or \
        "핥쨕" in message.content or \
        "햘짝" in message.content or \
        "햘쨕" in message.content or \
        "햝짝" in message.content or \
        "햝쨕" in message.content or \
        "호로" in message.content or \
        "호로놈" in message.content or \
        "호로새끼" in message.content or \
        "호로색" in message.content or \
        "호로쉑" in message.content or \
        "호로스까이" in message.content or \
        "호로스키" in message.content or \
        "후라들" in message.content or \
        "후래자식" in message.content or \
        "후레" in message.content or \
        "후뢰" in message.content:
            file = openpyxl.load_workbook("warning_list.xlsx")
            sheet = file.active
            member = discord.utils.get(client.get_all_members())
            for i in range(1, 129):
                if str(sheet["B" + str(i)].value) == str(message.author.id):
                    sheet["A" + str(i)].value = str(message.author.name)
                    warning_times = str(sheet["C" + str(i)].value)
                    break
                if str(sheet["B" + str(i)].value) == "-":
                    sheet["A" + str(i)].value = str(message.author.name)
                    sheet["B" + str(i)].value = str(message.author.id)
                    sheet["C" + str(i)].value = 0
                    warning_times = str(sheet["C" + str(i)].value)
                    break
            file.save("warning_list.xlsx")
            embed = discord.Embed(title="경고! 금지어 사용 감지됨." , description="**" + message.author.name + " | <@" + message.author.id + ">(이)가\n#" + message.channel.name + " | <#" + message.channel.id + ">에서 금지어 사용.\n현재 (" + str(warning_times) + "회 경고.)**\n경고는 <@150577293981515776>의 확인 이후에 누적 또는 기각됩니다.\n\n" + message.content, color=0xff0000)
            embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
            add_warning = await client.send_message(client.get_channel('542913851855994890'), embed=embed)
            await client.send_message(client.get_channel('539721276227584000'), embed=embed)
            await client.add_reaction(message, '⚠')
            await client.add_reaction(add_warning, '⚠')
            await client.add_reaction(add_warning, '🔶')

            #file = openpyxl.load_workbook("warning_list.xlsx")
            #sheet = file.active
            #member = discord.utils.get(client.get_all_members())
            #for i in range(1, 129):
            #    if str(sheet["B" + str(i)].value) == str(message.author.id):
            #        sheet["C" + str(i)].value = int(sheet["C" + str(i)].value) + 1
            #        warning_times = str(sheet["B" + str(i)].value)
            #        #if int(sheet["B" + str(i)].value) == 3:
            #            #await client.ban(member, 1)
            #        break
            #    if str(sheet["B" + str(i)].value) == "-":
            #        sheet["A" + str(i)].value = str(message.author.name)
            #        sheet["B" + str(i)].value = str(message.author.id)
            #        sheet["C" + str(i)].value = 1
            #        warning_times = str(sheet["C" + str(i)].value)
            #        break
            #file.save("warning_list.xlsx")
            #embed = discord.Embed(title="경고! 금지어 사용 감지됨." , description="**" + message.author.name + " | <@" + message.author.id + ">(이)가\n#" + message.channel.name + " | <#" + message.channel.id + ">에서 금지어 사용.\n(" + str(warning_times) + "회 경고.)**\n\n" + message.content, color=0xff0000)
            #embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
            #await client.send_message(client.get_channel('542913851855994890'), embed=embed)
            #await client.send_message(client.get_channel('539721276227584000'), embed=embed)
            #await client.send_message(message.channel, embed=embed)

access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
